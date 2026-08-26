"""
Leitor da lista de fornecedores (grupo de contas 2.1.3.01) do Plano de
Contas da Haroke. Mesmo leiaute observado na Antoninho (código reduzido na
coluna 0, classificação "2.1.3.01.xxxx" na coluna 7), mas aqui o nome do
fornecedor é procurado numa faixa de colunas (11 a 20) em vez de uma coluna
fixa, porque o export da Haroke não alinha o nome sempre na mesma posição.
"""
import openpyxl

from core.antoninho.classify import strip_accents


def norm(s: str) -> str:
    s = strip_accents(str(s)).upper()
    import re
    s = re.sub(r'[^A-Z0-9 ]', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()


def load_fornecedores(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    accounts = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        code = row[0].value
        classif = row[7].value if len(row) > 7 else None
        if code is None or classif is None:
            continue
        nome = None
        for idx in range(11, min(21, len(row))):
            val = row[idx].value
            if val is not None and not isinstance(val, float):
                nome = val
        if nome is None:
            continue
        if isinstance(classif, str) and classif.startswith('2.1.3.01'):
            accounts.append({'code': str(code), 'nome': str(nome).strip(), 'norm': norm(nome)})
    return accounts
