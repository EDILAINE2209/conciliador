"""
Leitor genérico da lista de fornecedores (grupo de contas 2.1.3.01) de um
Plano de Contas exportado do sistema contábil. Compartilhado entre as
empresas (Antoninho, Haroke, e as próximas) porque o leiaute observado até
agora é o mesmo nas duas: código reduzido na coluna 0, classificação
"2.1.3.01.xxxx" na coluna 7, e o nome do fornecedor em alguma coluna entre
11 e 20 (não é sempre a mesma posição — por isso a varredura em vez de uma
coluna fixa).

Movido para cá (fora de core/haroke/) porque deixou de ser algo específico
da Haroke assim que a Antoninho passou a usar o mesmo leitor para sugerir
conta de fornecedores novos a partir de um Plano de Contas.
"""
import re

import openpyxl

from core.antoninho.classify import strip_accents


def norm(s: str) -> str:
    s = strip_accents(str(s)).upper()
    s = re.sub(r'[^A-Z0-9 ]', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()


def load_fornecedores(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    accounts = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        code = row[0].value
        classif = row[7].value if len(row) > 7 else None
        if code is None or classif is None:
            continue
        nome = None
        for idx in range(11, min(21, len(row))):
            val = row[idx].value
            if val is not None and not isinstance(val, float):
                nome = val
        if nome is None:
            continue
        if isinstance(classif, str) and classif.startswith('2.1.3.01'):
            accounts.append({'code': str(code), 'nome': str(nome).strip(), 'norm': norm(nome)})
    return accounts
